from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from typing import List, Dict


class ExcelReportGenerator:
    """Generate Excel reports using openpyxl"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=16)
        self.pass_fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_course_grade_report(self, course_data: Dict, students_data: List[Dict], 
                                   statistics: Dict) -> bytes:
        """Generate course grade report Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Calificaciones"
        
        # Title
        ws['A1'] = f"Reporte de Calificaciones - {course_data['nombre']}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:F1')
        
        # Course Info
        ws['A3'] = "Código:"
        ws['B3'] = course_data['codigo']
        ws['D3'] = "Fecha:"
        ws['E3'] = datetime.now().strftime('%d/%m/%Y')
        
        ws['A4'] = "Total Estudiantes:"
        ws['B4'] = len(students_data)
        ws['D4'] = "Nota Aprobación:"
        ws['E4'] = statistics.get('nota_aprobacion', 60)
        
        # Statistics
        row = 6
        ws[f'A{row}'] = "ESTADÍSTICAS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        stats_data = [
            ('Promedio General', statistics.get('promedio', 0)),
            ('Nota Más Alta', statistics.get('nota_maxima', 0)),
            ('Nota Más Baja', statistics.get('nota_minima', 0)),
            ('Estudiantes Aprobados', f"{statistics.get('aprobados', 0)} ({statistics.get('tasa_aprobacion', 0):.1f}%)"),
            ('Estudiantes Reprobados', statistics.get('reprobados', 0)),
        ]
        
        for label, value in stats_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Student Grades
        row += 2
        ws[f'A{row}'] = "CALIFICACIONES POR ESTUDIANTE"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        header_row = row
        
        # Headers
        max_parciales = max([len(s.get('notas_parciales', [])) for s in students_data], default=0)
        headers = ['#', 'Estudiante', 'Email']
        for i in range(1, max_parciales + 1):
            headers.append(f'Parcial {i}')
        headers.extend(['Nota Final', 'Estado'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Data
        for idx, student in enumerate(students_data, 1):
            row += 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=student.get('nombre', ''))
            ws.cell(row=row, column=3, value=student.get('email', ''))
            
            notas = student.get('notas_parciales', [])
            for i in range(max_parciales):
                nota = notas[i] if i < len(notas) else '-'
                cell = ws.cell(row=row, column=4+i, value=nota if nota == '-' else float(nota))
                if nota != '-':
                    cell.number_format = '0.00'
            
            final_col = 4 + max_parciales
            ws.cell(row=row, column=final_col, value=student.get('nota_final', 0))
            ws.cell(row=row, column=final_col).number_format = '0.00'
            
            estado_cell = ws.cell(row=row, column=final_col+1, value=student.get('estado', 'N/A'))
            
            # Color row based on pass/fail
            fill = self.pass_fill if student.get('estado') == 'APROBADO' else self.fail_fill
            for col in range(1, final_col + 2):
                ws.cell(row=row, column=col).fill = fill
                ws.cell(row=row, column=col).border = self.border
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def create_student_performance_report(self, student_data: Dict, courses_data: List[Dict]) -> bytes:
        """Generate individual student performance report Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Desempeño"
        
        # Title
        ws['A1'] = f"Reporte de Desempeño - {student_data['nombre']} {student_data.get('apellido', '')}"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        # Student Info
        ws['A3'] = "Email:"
        ws['B3'] = student_data.get('email', 'N/A')
        ws['D3'] = "Fecha:"
        ws['E3'] = datetime.now().strftime('%d/%m/%Y')
        
        ws['A4'] = "Total Cursos:"
        ws['B4'] = len(courses_data)
        
        # Courses
        row = 6
        for course in courses_data:
            ws[f'A{row}'] = course['nombre']
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws[f'A{row}'] = "Parcial"
            ws[f'B{row}'] = "Nota"
            ws[f'C{row}'] = "Peso"
            ws[f'D{row}'] = "Contribución"
            
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
            
            parciales = course.get('parciales', [])
            for p in parciales:
                row += 1
                ws[f'A{row}'] = p.get('nombre', f"Parcial {p['numero']}")
                ws[f'B{row}'] = p['nota']
                ws[f'B{row}'].number_format = '0.00'
                ws[f'C{row}'] = f"{p['peso']}%"
                ws[f'D{row}'] = p['nota'] * p['peso'] / 100
                ws[f'D{row}'].number_format = '0.00'
            
            row += 1
            ws[f'C{row}'] = "Total:"
            ws[f'C{row}'].font = Font(bold=True)
            ws[f'D{row}'] = course['nota_final']
            ws[f'D{row}'].number_format = '0.00'
            ws[f'D{row}'].font = Font(bold=True)
            
            row += 1
            ws[f'A{row}'] = "Estado:"
            estado_cell = ws[f'B{row}']
            estado_cell.value = course['estado']
            estado_cell.font = Font(bold=True, color="008000" if course['estado'] == 'APROBADO' else "FF0000")
            
            row += 2
        
        # Summary
        aprobados = sum(1 for c in courses_data if c['estado'] == 'APROBADO')
        promedio_general = sum(c['nota_final'] for c in courses_data) / len(courses_data) if courses_data else 0
        
        ws[f'A{row}'] = "RESUMEN GENERAL"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = "Cursos Aprobados:"
        ws[f'B{row}'] = f"{aprobados}/{len(courses_data)}"
        row += 1
        ws[f'A{row}'] = "Promedio General:"
        ws[f'B{row}'] = promedio_general
        ws[f'B{row}'].number_format = '0.00'
        
        # Auto-adjust columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def create_system_overview_report(self, overview_data: Dict) -> bytes:
        """Generate system overview report Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen Sistema"
        
        # Title
        ws['A1'] = "Reporte General del Sistema"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # System Stats
        row = 3
        ws[f'A{row}'] = "ESTADÍSTICAS GENERALES"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 1
        stats_data = [
            ('Total Cursos', overview_data.get('total_cursos', 0)),
            ('Total Estudiantes', overview_data.get('total_estudiantes', 0)),
            ('Total Profesores', overview_data.get('total_profesores', 0)),
            ('Promedio General', overview_data.get('promedio_general', 0)),
            ('Tasa de Aprobación', f"{overview_data.get('tasa_aprobacion', 0):.1f}%"),
        ]
        
        for label, value in stats_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Courses Performance
        if 'cursos_performance' in overview_data:
            row += 2
            ws[f'A{row}'] = "DESEMPEÑO POR CURSO"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            
            row += 1
            headers = ['Curso', 'Estudiantes', 'Promedio', 'Aprobación %']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
            
            for curso in overview_data['cursos_performance']:
                row += 1
                ws.cell(row=row, column=1, value=curso['nombre'])
                ws.cell(row=row, column=2, value=curso['total_estudiantes'])
                cell = ws.cell(row=row, column=3, value=curso['promedio'])
                cell.number_format = '0.00'
                ws.cell(row=row, column=4, value=f"{curso['tasa_aprobacion']:.1f}%")
        
        # Auto-adjust columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
